import sys
import time
sys.path.append(sys.path[0] + "\\site-packages")
from watchdog.events import LoggingEventHandler
from watchdog.observers import Observer
from importlib import reload
# import your_sample_module_file chanllenge
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook(write_only=True)
ws = wb.create_sheet()

print('Monitoring Started')

class SaveDataToExcel(LoggingEventHandler):
    def on_modified(self, event):
        try:
            wb1 = load_workbook('1.xlsx')
            ws1 = wb1.active
            ws.append([value.value for l in ws1['A2':'T2'] for value in l])
            print("Data Appended")
            restart_server = True
        except (FileNotFoundError, PermissionError):
            pass 

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else '.'
    observer = Observer()
    observer.schedule(SaveDataToExcel(), path, recursive=True)

    observer.start()
    try:
        while True:         
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        wb.save('dataraw.xlsx')
    observer.join()